import io
import os
from datetime import datetime
from typing import Optional, List, Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.config import (
    EXPORTS_DIR, CURRENCY_SYMBOL,
    GROUP_ACER_MONITORS, GROUP_OTHER_PRODUCTS, GROUP_ALL,
    EXCEL_MONITORS_FILENAME, EXCEL_OTHER_FILENAME, EXCEL_ALL_FILENAME
)
from app.database import (
    get_all_products, get_products_by_group,
    get_price_history_for_asin, get_product_statistics
)
from app.history_engine import get_22_month_labels

def build_excel_workbook(
    products: List[Dict[str, Any]],
    report_title: str = "ACER AMAZON PRICE INTELLIGENCE REPORT",
    group_label: str = "All Products"
) -> openpyxl.Workbook:
    """
    Generates a structured, executive-styled monotone Excel workbook containing:
    1. Product_Overview: Specs, live prices, discounts, ATL/ATH metrics, clickable Amazon links.
    2. 6_Months_Price_History: Full monthly matrix across all 6 months with direct links and deal markers.
    3. Monthly_Statistics: Category-level and portfolio aggregate metrics over 6 months.
    """
    wb = openpyxl.Workbook()
    
    # Executive Monotone Typography & Palette
    FONT_FAMILY = "Segoe UI"
    
    title_font = Font(name=FONT_FAMILY, size=15, bold=True, color="FFFFFF")
    subtitle_font = Font(name=FONT_FAMILY, size=9, italic=True, color="CBD5E1")
    header_font = Font(name=FONT_FAMILY, size=10, bold=True, color="FFFFFF")
    sub_header_font = Font(name=FONT_FAMILY, size=9, bold=True, color="FFFFFF")
    data_font = Font(name=FONT_FAMILY, size=9, color="1E293B")
    bold_data_font = Font(name=FONT_FAMILY, size=9, bold=True, color="0F172A")
    link_font = Font(name=FONT_FAMILY, size=9, color="2563EB", underline="single")
    link_bold_font = Font(name=FONT_FAMILY, size=9, bold=True, color="1E293B", underline="single")
    badge_green_font = Font(name=FONT_FAMILY, size=9, bold=True, color="166534")
    
    # Monotone Enterprise Fills
    title_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid") # Obsidian Slate
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Dark Charcoal
    sub_header_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid") # Slate
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid") # Off-white silver
    best_deal_fill = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid") # Subdued Mint
    out_of_stock_fill = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid") # Subdued Rose
    
    thin_border_side = Side(style='thin', color='CBD5E1')
    border_all = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_wrap_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    month_labels = get_22_month_labels()

    # =========================================================================
    # SHEET 1: Product_Overview
    # =========================================================================
    ws1 = wb.active
    ws1.title = "Product_Overview"
    ws1.views.sheetView[0].showGridLines = True

    # Title Banner
    ws1.merge_cells("A1:M1")
    title_cell = ws1["A1"]
    title_cell.value = report_title.upper()
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws1.row_dimensions[1].height = 36

    ws1.merge_cells("A2:M2")
    sub_cell = ws1["A2"]
    sub_cell.value = f"Dashboard: {group_label} | Generated on: {datetime.now().strftime('%d %B %Y, %I:%M %p')} | Active Tracked ASINs: {len(products)} | Currency: {CURRENCY_SYMBOL} (INR)"
    sub_cell.font = subtitle_font
    sub_cell.fill = title_fill
    sub_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws1.row_dimensions[2].height = 20

    # Headers for Sheet 1
    headers1 = [
        "ASIN", "Product Title", "Category", f"Today's Price ({CURRENCY_SYMBOL})",
        f"MRP ({CURRENCY_SYMBOL})", "Discount %", "Stock Status",
        f"22-Mo Lowest ({CURRENCY_SYMBOL})", f"22-Mo Highest ({CURRENCY_SYMBOL})",
        f"22-Mo Avg ({CURRENCY_SYMBOL})", "% Off ATH", "Rating", "Amazon Link"
    ]
    
    ws1.append([]) # Row 3 spacer
    ws1.row_dimensions[3].height = 8
    
    ws1.append(headers1) # Row 4
    ws1.row_dimensions[4].height = 26
    
    for col_idx, col_name in enumerate(headers1, 1):
        c = ws1.cell(row=4, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = align_center
        c.border = border_all

    # Data Rows for Sheet 1
    for r_idx, prod in enumerate(products, start=5):
        asin = prod["asin"]
        stats = get_product_statistics(asin)
        
        is_zebra = (r_idx % 2 == 0)
        row_fill = zebra_fill if is_zebra else None
        
        current_price = prod["current_price"]
        mrp = prod["mrp"]
        discount_pct = stats.get("discount_from_mrp", 0.0) / 100.0
        stock = prod["stock_status"]
        min_price = stats.get("min_price", current_price)
        max_price = stats.get("max_price", current_price)
        avg_price = stats.get("avg_price", current_price)
        off_ath = stats.get("discount_from_ath", 0.0) / 100.0
        rating = prod["rating"]
        url = prod["url"]

        row_values = [
            asin,
            prod["title"],
            prod["category"],
            current_price,
            mrp,
            discount_pct,
            stock,
            min_price,
            max_price,
            avg_price,
            off_ath,
            rating,
            "View on Amazon"
        ]
        ws1.append(row_values)
        ws1.row_dimensions[r_idx].height = 22

        # Style each cell in row
        for col_idx in range(1, len(row_values) + 1):
            cell = ws1.cell(row=r_idx, column=col_idx)
            cell.font = data_font
            cell.border = border_all
            if row_fill:
                cell.fill = row_fill

            # Column specific formatters
            if col_idx == 1: # ASIN
                cell.value = asin
                cell.hyperlink = url
                cell.alignment = align_center
                cell.font = link_bold_font
            elif col_idx == 2: # Title
                cell.alignment = align_wrap_left
            elif col_idx == 3: # Category
                cell.alignment = align_center
            elif col_idx in (4, 5, 8, 9, 10): # Currency amounts
                cell.number_format = f'{CURRENCY_SYMBOL}#,##0'
                cell.alignment = align_right
                if col_idx == 4 and stats.get("is_atl"):
                    cell.fill = best_deal_fill
                    cell.font = bold_data_font
            elif col_idx in (6, 11): # Percentages
                cell.number_format = '0.0%'
                cell.alignment = align_right
            elif col_idx == 7: # Stock
                cell.alignment = align_center
                if "out" in stock.lower():
                    cell.fill = out_of_stock_fill
            elif col_idx == 12: # Rating
                cell.number_format = '0.0'
                cell.alignment = align_center
            elif col_idx == 13: # Link
                cell.value = "Amazon ↗"
                cell.hyperlink = url
                cell.alignment = align_center
                cell.font = link_font

    # =========================================================================
    # SHEET 2: 6_Months_Price_History
    # =========================================================================
    ws2 = wb.create_sheet(title="6_Months_Price_History")
    ws2.views.sheetView[0].showGridLines = True

    total_h_cols = 4 + len(month_labels) + 2
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_h_cols)
    title_cell2 = ws2["A1"]
    title_cell2.value = f"{report_title} — 6-MONTH HISTORICAL PRICE MATRIX"
    title_cell2.font = title_font
    title_cell2.fill = title_fill
    title_cell2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws2.row_dimensions[1].height = 36

    ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_h_cols)
    sub_cell2 = ws2["A2"]
    sub_cell2.value = f"Complete monthly trajectory from {month_labels[0]} to {month_labels[-1]} | Includes seasonal sales"
    sub_cell2.font = subtitle_font
    sub_cell2.fill = title_fill
    sub_cell2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws2.row_dimensions[2].height = 20

    ws2.append([]) # Row 3 spacer
    ws2.row_dimensions[3].height = 8

    headers2 = ["ASIN", "Product Title", "Category", "Baseline MRP"] + month_labels + ["6-Mo Min", "6-Mo Max"]
    ws2.append(headers2)
    ws2.row_dimensions[4].height = 26

    for col_idx, col_name in enumerate(headers2, 1):
        c = ws2.cell(row=4, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = align_center
        c.border = border_all

    for r_idx, prod in enumerate(products, start=5):
        asin = prod["asin"]
        history = get_price_history_for_asin(asin)
        history_map = {h["month_label"]: h["price"] for h in history}
        
        prices_list = [history_map.get(m, prod["current_price"]) for m in month_labels]
        min_p = min(prices_list) if prices_list else prod["current_price"]
        max_p = max(prices_list) if prices_list else prod["current_price"]
        
        is_zebra = (r_idx % 2 == 0)
        row_fill = zebra_fill if is_zebra else None

        row_vals = [
            asin,
            prod["title"],
            prod["category"],
            prod["mrp"]
        ] + prices_list + [min_p, max_p]
        
        ws2.append(row_vals)
        ws2.row_dimensions[r_idx].height = 22

        for col_idx in range(1, len(row_vals) + 1):
            cell = ws2.cell(row=r_idx, column=col_idx)
            cell.font = data_font
            cell.border = border_all
            if row_fill:
                cell.fill = row_fill

            if col_idx == 1:
                cell.value = asin
                cell.hyperlink = prod["url"]
                cell.alignment = align_center
                cell.font = link_bold_font
            elif col_idx == 2:
                cell.alignment = align_wrap_left
            elif col_idx == 3:
                cell.alignment = align_center
            elif col_idx >= 4:
                cell.number_format = f'{CURRENCY_SYMBOL}#,##0'
                cell.alignment = align_right
                val = cell.value
                if isinstance(val, (int, float)) and val == min_p:
                    cell.fill = best_deal_fill # Highlight monthly lowest price in timeline

    # =========================================================================
    # SHEET 3: Monthly_Statistics
    # =========================================================================
    ws3 = wb.create_sheet(title="Monthly_Statistics")
    ws3.views.sheetView[0].showGridLines = True

    total_s_cols = 2 + len(month_labels)
    ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_s_cols)
    title_cell3 = ws3["A1"]
    title_cell3.value = f"{report_title} — CATEGORY & PORTFOLIO MONTHLY BENCHMARKS"
    title_cell3.font = title_font
    title_cell3.fill = title_fill
    title_cell3.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws3.row_dimensions[1].height = 36

    ws3.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_s_cols)
    sub_cell3 = ws3["A2"]
    sub_cell3.value = f"Average Category Price Trends across 6-Month Observation Period ({CURRENCY_SYMBOL} INR)"
    sub_cell3.font = subtitle_font
    sub_cell3.fill = title_fill
    sub_cell3.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws3.row_dimensions[2].height = 20

    ws3.append([]) # Row 3 spacer
    ws3.row_dimensions[3].height = 8

    headers3 = ["Category / Scope", "Metric"] + month_labels
    ws3.append(headers3)
    ws3.row_dimensions[4].height = 26

    for col_idx, col_name in enumerate(headers3, 1):
        c = ws3.cell(row=4, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = align_center
        c.border = border_all

    # Category monthly averages
    categories = sorted(list(set(p["category"] for p in products))) if products else []
    cat_prods = {cat: [p for p in products if p["category"] == cat] for cat in categories}
    
    current_r = 5
    for cat in categories:
        prods_in_cat = cat_prods[cat]
        cat_month_totals = {m: [] for m in month_labels}
        
        for p in prods_in_cat:
            h_list = get_price_history_for_asin(p["asin"])
            h_dict = {h["month_label"]: h["price"] for h in h_list}
            for m in month_labels:
                if m in h_dict:
                    cat_month_totals[m].append(h_dict[m])

        cat_avgs = [
            (sum(cat_month_totals[m]) / len(cat_month_totals[m])) if cat_month_totals[m] else 0.0
            for m in month_labels
        ]

        row_vals = [cat, f"Avg Price ({len(prods_in_cat)} items)"] + cat_avgs
        ws3.append(row_vals)
        ws3.row_dimensions[current_r].height = 22
        
        for col_idx in range(1, len(row_vals) + 1):
            cell = ws3.cell(row=current_r, column=col_idx)
            cell.font = data_font
            cell.border = border_all
            if col_idx == 1:
                cell.alignment = align_left
                cell.font = bold_data_font
            elif col_idx == 2:
                cell.alignment = align_center
            else:
                cell.number_format = f'{CURRENCY_SYMBOL}#,##0'
                cell.alignment = align_right
        current_r += 1

    # Portfolio Total Row
    if products:
        all_month_totals = {m: [] for m in month_labels}
        for p in products:
            h_list = get_price_history_for_asin(p["asin"])
            h_dict = {h["month_label"]: h["price"] for h in h_list}
            for m in month_labels:
                if m in h_dict:
                    all_month_totals[m].append(h_dict[m])

        portfolio_avgs = [
            (sum(all_month_totals[m]) / len(all_month_totals[m])) if all_month_totals[m] else 0.0
            for m in month_labels
        ]
        
        ws3.append([]) # spacer
        current_r += 1
        
        row_vals = ["Overall Portfolio Benchmark", f"Weighted Average ({len(products)} items)"] + portfolio_avgs
        ws3.append(row_vals)
        ws3.row_dimensions[current_r].height = 24
        
        for col_idx in range(1, len(row_vals) + 1):
            cell = ws3.cell(row=current_r, column=col_idx)
            cell.font = bold_data_font
            cell.fill = sub_header_fill
            cell.font = sub_header_font
            cell.border = border_all
            if col_idx == 1:
                cell.alignment = align_left
            elif col_idx == 2:
                cell.alignment = align_center
            else:
                cell.number_format = f'{CURRENCY_SYMBOL}#,##0'
                cell.alignment = align_right

    # =========================================================================
    # Auto-Fit Column Widths for all sheets
    # =========================================================================
    for sheet in [ws1, ws2, ws3]:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row in (1, 2):
                    continue
                if cell.value:
                    val_str = str(cell.value)
                    if len(val_str) > max_len:
                        max_len = len(val_str)
            
            # Constrain min and max column widths
            if col_letter == 'A':
                sheet.column_dimensions[col_letter].width = 16
            elif col_letter == 'B':
                sheet.column_dimensions[col_letter].width = 48
            else:
                sheet.column_dimensions[col_letter].width = max(13, min(max_len + 3, 26))

    return wb

def export_monitors_excel(target_path: Optional[str] = None) -> str:
    """Generates and saves the dedicated Acer Monitors Excel report."""
    products = get_products_by_group(GROUP_ACER_MONITORS)
    wb = build_excel_workbook(
        products=products,
        report_title="Acer Monitors Amazon Price Intelligence Report",
        group_label="Acer Monitors"
    )
    if not target_path:
        target_path = str(EXPORTS_DIR / EXCEL_MONITORS_FILENAME)
    wb.save(target_path)
    return target_path

def export_other_products_excel(target_path: Optional[str] = None) -> str:
    """Generates and saves the dedicated Other Products Excel report."""
    products = get_products_by_group(GROUP_OTHER_PRODUCTS)
    wb = build_excel_workbook(
        products=products,
        report_title="Amazon Other Products Price Intelligence Report",
        group_label="Other Products"
    )
    if not target_path:
        target_path = str(EXPORTS_DIR / EXCEL_OTHER_FILENAME)
    wb.save(target_path)
    return target_path

def export_all_portfolio_excel(target_path: Optional[str] = None) -> str:
    """Generates and saves the combined All Products Excel report."""
    products = get_all_products()
    wb = build_excel_workbook(
        products=products,
        report_title="Acer Amazon Portfolio Price Intelligence Report",
        group_label="All Tracked Products"
    )
    if not target_path:
        target_path = str(EXPORTS_DIR / EXCEL_ALL_FILENAME)
    wb.save(target_path)
    return target_path

def export_excel_by_group(group: str, target_path: Optional[str] = None) -> str:
    """Routes export to the appropriate group workbook."""
    grp = group.lower() if group else GROUP_ALL
    if grp == GROUP_ACER_MONITORS:
        return export_monitors_excel(target_path)
    elif grp == GROUP_OTHER_PRODUCTS:
        return export_other_products_excel(target_path)
    else:
        return export_all_portfolio_excel(target_path)

def export_excel_to_file(target_path: Optional[str] = None) -> str:
    """Default backward-compatible helper exporting all products."""
    return export_all_portfolio_excel(target_path)

def export_excel_to_bytes(group: str = GROUP_ALL) -> io.BytesIO:
    """Returns Excel file as an in-memory BytesIO stream."""
    if group == GROUP_ACER_MONITORS:
        products = get_products_by_group(GROUP_ACER_MONITORS)
        title = "Acer Monitors Amazon Price Intelligence Report"
        label = "Acer Monitors"
    elif group == GROUP_OTHER_PRODUCTS:
        products = get_products_by_group(GROUP_OTHER_PRODUCTS)
        title = "Amazon Other Products Price Intelligence Report"
        label = "Other Products"
    else:
        products = get_all_products()
        title = "Acer Amazon Portfolio Price Intelligence Report"
        label = "All Products"
        
    wb = build_excel_workbook(products, title, label)
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
